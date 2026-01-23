import os
import json
import time
import hashlib
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook
from flask import Flask, request, jsonify, render_template, redirect, url_for, session, send_file

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "CHANGE_THIS_SECRET")

# ---------------- CONFIG ----------------
ADMIN_USER = os.environ.get("ADMIN_USER", "Mkloveinfinite@#")
ADMIN_PASS = os.environ.get("ADMIN_PASS", "Mkundefined@#")
MAX_DEVICES = int(os.environ.get("MAX_DEVICES", "5"))

DATA_FILE = "data.json"
REPORT_FILE = "work_report.xlsx"

# ---------------- HELPERS: GENERAL ----------------
def now_str():
    return datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")

def load_db():
    if not os.path.exists(DATA_FILE):
        return {"devices": [], "logs": []}
    with open(DATA_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def save_db(db):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(db, f, indent=2)

def hash_pw(pw: str):
    return hashlib.sha256(pw.encode("utf-8")).hexdigest()

def check_admin(username, password):
    return username == ADMIN_USER and password == ADMIN_PASS

def get_ip():
    ip = request.headers.get("X-Forwarded-For", "")
    if ip:
        return ip.split(",")[0].strip()
    return request.remote_addr

def log_event(db, device_id, name, status):
    db["logs"].insert(0, {
        "device_id": device_id,
        "name": name,
        "ip": get_ip(),
        "status": status,
        "time": now_str()
    })
    db["logs"] = db["logs"][:50]

def find_device(db, device_id):
    for d in db["devices"]:
        if d["device_id"] == device_id:
            return d
    return None

# ---------------- HELPERS: REPORT & EXCEL (NEW) ----------------
YOUTUBE_REGEX = re.compile(
    r"(https?://)?(www\.)?(youtube\.com/watch\?v=|youtu\.be/)[A-Za-z0-9_\-]{6,}"
)

def is_valid_youtube_link(text: str) -> bool:
    if not text:
        return False
    return bool(YOUTUBE_REGEX.search(text.strip()))

def ensure_excel_file():
    if os.path.exists(REPORT_FILE):
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "report"

    # Header Row
    ws.append([
        "time_utc",
        "username",
        "device_id",
        "device_name",
        "event",
        "valid_links",
        "invalid_text",
        "scrape_done",
        "download_done",
        "seconds"
    ])
    wb.save(REPORT_FILE)

def append_report_row(row: list):
    ensure_excel_file()
    try:
        wb = load_workbook(REPORT_FILE)
        ws = wb["report"]
        ws.append(row)
        wb.save(REPORT_FILE)
    except Exception as e:
        print(f"Error saving to Excel: {e}")

# ---------------- API: LOGIN & PING ----------------

@app.route("/api/login", methods=["POST"])
def api_login():
    data = request.get_json(force=True)

    username = data.get("username", "")
    password = data.get("password", "")
    device_id = data.get("device_id", "")
    device_name = data.get("device_name", "Unknown-PC")

    db = load_db()

    # check credentials
    if not check_admin(username, password):
        log_event(db, device_id, device_name, "INVALID_CREDENTIALS")
        save_db(db)
        return jsonify({"ok": False, "error": "Invalid username or password"}), 401

    if not device_id:
        log_event(db, device_id, device_name, "NO_DEVICE_ID")
        save_db(db)
        return jsonify({"ok": False, "error": "Device ID missing"}), 400

    existing = find_device(db, device_id)

    # if device exists but disabled
    if existing and existing.get("disabled", False):
        existing["last_seen"] = now_str()
        existing["ip"] = get_ip()
        log_event(db, device_id, device_name, "DEVICE_DISABLED")
        save_db(db)
        return jsonify({"ok": False, "error": "This device is disabled by admin"}), 403

    # if device not registered, register only if slots available
    if not existing:
        if len(db["devices"]) >= MAX_DEVICES:
            log_event(db, device_id, device_name, "MAX_DEVICES_REACHED")
            save_db(db)
            return jsonify({"ok": False, "error": "Max devices reached. Contact admin."}), 403

        new_device = {
            "id": int(time.time()),
            "device_id": device_id,
            "device_name": device_name,
            "ip": get_ip(),
            "created": now_str(),
            "last_seen": now_str(),
            "disabled": False
        }
        db["devices"].append(new_device)
        log_event(db, device_id, device_name, "DEVICE_REGISTERED")

    else:
        existing["device_name"] = device_name
        existing["ip"] = get_ip()
        existing["last_seen"] = now_str()
        log_event(db, device_id, device_name, "LOGIN_OK")

    save_db(db)
    return jsonify({"ok": True, "max_devices": MAX_DEVICES})

@app.route("/api/ping", methods=["GET"])
def api_ping():
    return jsonify({
        "ok": True,
        "message": "pong",
        "time": now_str()
    }), 200

# ---------------- API: REPORTING (LOG TO EXCEL) ----------------

@app.route("/api/report/paste_links", methods=["POST"])
def report_paste_links():
    data = request.get_json(force=True)

    username = data.get("username", "")
    device_id = data.get("device_id", "")
    device_name = data.get("device_name", "Unknown")
    items = data.get("items", [])

    valid_count = 0
    invalid_count = 0

    for x in items:
        if is_valid_youtube_link(str(x)):
            valid_count += 1
        else:
            invalid_count += 1

    append_report_row([
        now_str(),
        username,
        device_id,
        device_name,
        "paste_links",
        valid_count,
        invalid_count,
        "",
        "",
        ""
    ])

    return jsonify({"ok": True, "valid": valid_count, "invalid": invalid_count}), 200

@app.route("/api/report/scrape_done", methods=["POST"])
def report_scrape_done():
    data = request.get_json(force=True)
    username = data.get("username", "")
    device_id = data.get("device_id", "")
    device_name = data.get("device_name", "Unknown")
    count = int(data.get("count", 0))

    append_report_row([
        now_str(),
        username,
        device_id,
        device_name,
        "scrape_done",
        "",
        "",
        count,
        "",
        ""
    ])

    return jsonify({"ok": True}), 200

@app.route("/api/report/download_done", methods=["POST"])
def report_download_done():
    data = request.get_json(force=True)
    username = data.get("username", "")
    device_id = data.get("device_id", "")
    device_name = data.get("device_name", "Unknown")
    count = int(data.get("count", 0))

    append_report_row([
        now_str(),
        username,
        device_id,
        device_name,
        "download_done",
        "",
        "",
        "",
        count,
        ""
    ])

    return jsonify({"ok": True}), 200

@app.route("/api/report/session", methods=["POST"])
def report_session():
    data = request.get_json(force=True)
    username = data.get("username", "")
    device_id = data.get("device_id", "")
    device_name = data.get("device_name", "Unknown")
    seconds = int(data.get("seconds", 0))

    append_report_row([
        now_str(),
        username,
        device_id,
        device_name,
        "session_time",
        "",
        "",
        "",
        "",
        seconds
    ])

    return jsonify({"ok": True}), 200

# ---------------- ADMIN PANEL ----------------

@app.route("/")
def home():
    if not session.get("admin"):
        return redirect(url_for("admin_login"))
    db = load_db()
    return render_template("dashboard.html", devices=db["devices"], logs=db["logs"], max_devices=MAX_DEVICES)

@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        username = request.form.get("username", "")
        password = request.form.get("password", "")
        if check_admin(username, password):
            session["admin"] = True
            return redirect(url_for("home"))
        return render_template("login.html", error="Invalid Admin Login")
    return render_template("login.html", error=None)

@app.route("/admin/logout")
def admin_logout():
    session.clear()
    return redirect(url_for("admin_login"))

@app.route("/admin/device/remove/<device_id>", methods=["POST"])
def remove_device(device_id):
    if not session.get("admin"):
        return redirect(url_for("admin_login"))
    db = load_db()
    db["devices"] = [d for d in db["devices"] if d["device_id"] != device_id]
    save_db(db)
    return redirect(url_for("home"))

@app.route("/admin/device/disable/<device_id>", methods=["POST"])
def disable_device(device_id):
    if not session.get("admin"):
        return redirect(url_for("admin_login"))
    db = load_db()
    d = find_device(db, device_id)
    if d:
        d["disabled"] = True
    save_db(db)
    return redirect(url_for("home"))

@app.route("/admin/device/enable/<device_id>", methods=["POST"])
def enable_device(device_id):
    if not session.get("admin"):
        return redirect(url_for("admin_login"))
    db = load_db()
    d = find_device(db, device_id)
    if d:
        d["disabled"] = False
    save_db(db)
    return redirect(url_for("home"))

@app.route("/admin/work/excel")
def download_excel():
    if not session.get("admin"):
        return redirect(url_for("admin_login"))

    ensure_excel_file()
    return send_file(REPORT_FILE, as_attachment=True)

# health check for Render
@app.route("/health")
def health():
    return "OK", 200

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", "5000")))
